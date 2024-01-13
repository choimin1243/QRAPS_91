from openpyxl import Workbook  # 추가된 부분
import re
import pandas as pd
from io import BytesIO
from openpyxl.utils.dataframe import dataframe_to_rows
from fastapi.responses import StreamingResponse
from fastapi.templating import Jinja2Templates
from fastapi import FastAPI, Form, Request, Depends
from fastapi import APIRouter, Request
from database import SessionLocal  # SessionLocal을 불러옴
from rcpackage import hello2
from Lpackage import get_all_Lpackage_package_from_db, get_all_Lpackage_partnumber_from_db
from fastapi import UploadFile, File
from datetime import datetime
import openpyxl
import json
import os
import tempfile
import openai



db = SessionLocal()

appends = APIRouter()
templates = Jinja2Templates(directory="templates")
templates.env.globals.update(enumerate=enumerate)

original_list = get_all_Lpackage_partnumber_from_db(db)
hh1=get_all_Lpackage_package_from_db(db)


def flatten(lst):
    result = []
    for item in lst:
        if isinstance(item, list):
            result.extend(flatten(item))
        else:
            result.append(item)
    return result


def filterate(listly):
    database_list = hello2(db)
    listly = flatten(listly)
    list_result = ' '.join(map(str, listly))

    x = ""
    for i in range(len(database_list)):
        finder = database_list[i]
        if finder in list_result:
            x = finder
            break

    return x


# HTML 템플릿을 렌더링하는 엔드포인트
@appends.get("/")
async def render_upload_form(request: Request):

    return templates.TemplateResponse("index.html", {"request": request})


def get_Lpackage_filter_Lpack(listly):
    original_list=get_all_Lpackage_package_from_db(db)
    original_list = [s.strip() for s in original_list]

    listly=flatten(listly)
    listly=[str(i).strip() for i in listly]


    x=""
    for i in range(len(original_list)):
        finder=original_list[i]

        for s in range(len(listly)):
            if finder in listly[s]:
                x=finder
                break
    return x


def get_Lpartnmuber_filter_Lpack(listly):
    original_list=get_all_Lpackage_partnumber_from_db(db)

    original_list = [s.strip() for s in original_list]

    listly=flatten(listly)
    listly=[str(i).strip() for i in listly]


    x=""
    for i in range(len(original_list)):
        finder=original_list[i]

        for s in range(len(listly)):
            if finder in listly[s]:
                x=finder
                break
    return x


def wat_resize(wat):
    패턴 = r"(\d+)/(\d+)"
    매치 = re.search(패턴, wat)
    pp = r'\b(\d+)w\b'
    pattern = r'\b(\d+)\s*(?:밀리와트|mW|mw)\b'
    matcheses = re.findall(pattern, wat, re.IGNORECASE)

    패턴즈 = r"kw"
    m = r'mw'
    매치들 = re.findall(패턴즈, wat, re.IGNORECASE)
    if 매치들:
        if 매치:
            분자 = int(매치.group(1))
            분모 = int(매치.group(2))
            소수 = 분자 / 분모
            결과_문자열 = float(소수) * 1000

            return 결과_문자열

        else:
            패턴 = r"\d+"
            추출된_숫자들 = re.findall(패턴, wat, re.IGNORECASE)
            return float(추출된_숫자들[0]) * 1000

    와트매치 = re.findall(pp, wat, re.IGNORECASE)
    if 와트매치:
        # print(wat, "@@)")
        if 매치:
            분자 = int(매치.group(1))
            분모 = int(매치.group(2))
            소수 = 분자 / 분모
            결과_문자열 = float(소수)

            return 결과_문자열

        else:
            패턴 = r"\d+"
            추출된_숫자들 = re.findall(패턴, wat, re.IGNORECASE)
            return 추출된_숫자들[0]
    m = r'mw'
    MW = re.findall(m, wat, re.IGNORECASE)

    if MW:
        if 매치:
            분자 = int(매치.group(1))
            분모 = int(매치.group(2))
            소수 = 분자 / 분모
            return float(소수) * 0.001

        else:
            패턴 = r"\d+"
            추출된_숫자들 = re.findall(패턴, wat, re.IGNORECASE)
            return float(추출된_숫자들[0]) * 0.001

def extract_number(text):
    match = re.search(r'(\d+)(K|k)', text)
    if match:
        return match.group(1)
    else:
        return None





def om(rest):
    pattern = r"(?<!\w)(\d+)\s*Ω"
    patternk = r'(\d+\.\d+|\d+)(?=\s*[KΩ|㏀])'
    patternm = r'\d+\s*m[Ω|Ω]'
    patternss = r'\d+\s*[^mM]㏁|\d+\s*MΩ'
    pattern_float=r'\d+\.\d+\s*Ω'
    if re.search(pattern_float,rest) or re.search(pattern, rest):
        patternr = r"\d+"
        return float(re.findall(patternr, rest)[0]) * 0.001


    if re.search(patternk, rest):
        return float(re.findall(patternk, rest)[0])
    if re.search(patternm, rest):
        patternr = r"\d+"
        return float(re.findall(patternr, rest)[0]) * 0.000001
    if re.search(patternss, rest):
        patternr = r"\d+"
        return float(re.findall(patternr, rest)[0]) * 1000





@appends.post("/send-list/")
async def send_list(request: Request, selected_columns: str = Form(...), content_items: str = Form(...),
                    encoded_data: str = Form(...)):
    selected_columns = json.loads(selected_columns)
    content_items = json.loads(content_items)
    encoded_data = json.loads(encoded_data)

    number_to = len(content_items)
    output_excel = BytesIO()
    work = Workbook()

    for charact in content_items:
        data_list = encoded_data  # JSON 문자열을 리스트로 변환
        work.create_sheet(title=charact)

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # 데이터를 시트에 추가
        for row_data in data_list:
            sheet.append(row_data)

        last_column = sheet.max_column
        last_row = sheet.max_row

        data = []
        location_column_index = selected_columns
        part_number=None
        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=last_row, max_col=last_column):
            # 한 행의 데이터를 저장할 리스트를 생성합니다.
            row_data = []
            for idx, cell in enumerate(row, 1):
                row_data.append([cell.value])
            data.append(row_data)

        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=last_row, max_col=last_column):
            # 한 행의 데이터를 저장할 리스트를 생성합니다.
            for idx, cell in enumerate(row, 1):
                if cell.value is not None and isinstance(cell.value, str):
                    cell.value = cell.value.lower()
                    if cell.value == "package":
                        part_number = idx

        if part_number == None:
            part_number = 500

        # print(part_number,"@@@~~~~~")
        def remove_duplicates(input_list):
            return list(set(input_list))

        patternwat = r"(\d+(?:/\d+)?(?:\.\d+)?)\s*(w|kw|mw)"
        patternnp = r"(?<!\w)(?<!\d)\d+(?:\.\d+)?(?:\s*(?:pF|nF|uF|µF|UF|p|n|u|µ))(?!\w)"
        patternv = r"(\d+(?:\.\d+)?)\s*(?:[kK]?[mM]?[vV])"

        tolerance_values = ["J", "F", "A", "B", "G", "M", "Z"]


        # 수정된 정규식
        pattern_tor = r"(?<![A-Za-z0-9.,-])(?:{})(?![A-Za-z0-9.,])".format("|".join(tolerance_values))

        patternAed = r"([-+]?\d+(?:\.\d+)?)\s*([mµ]?[AaKk])"
        patterntemp = r'\d+(?:\.\d+)?\s*℃'

        location_column_index = remove_duplicates(location_column_index)
        pattern_kv = r"kv"
        pattern_v = r"\d+(?:\.\d+)?(?=\s*(?i)v)"

        result_data = []

        DIOD=[r"(?<!\S)"+"D"+"(\d+)",r"(?<!\S)"+"LED"+"(\d+)",r"(?<!\S)"+"ZD"+"(\d+)",r"(?<!\S)"+"BD"+"(\d+)"]


        character = charact
        pattern = r"(?<!\S)" + character + "(\d+)"

        list_row = []

        for i in range(len(data)):
            try:
                datas = data[i][int(location_column_index[0])]
                if datas[0] != None:
                    parsed_data = [row.replace(" ", "").split(",") for row in datas[0].split("\n")]
                    flattened_data = [item for sublist in parsed_data for item in sublist]
                    if re.findall(pattern, flattened_data[0], re.IGNORECASE) and character!="D":
                        list_row.append(i)
                        for s in range(len(flattened_data)):
                            if flattened_data[s] != '':
                                result_data.append([i, flattened_data[s].strip()])


                    if character=="D":
                        for k in range(len(DIOD)):
                            if re.findall(DIOD[k],flattened_data[0],re.IGNORECASE):
                                list_row.append(i)
                                for s in range(len(flattened_data)):
                                    if flattened_data[s] != '':
                                        result_data.append([i, flattened_data[s].strip()])



            except:
                pass

        voltage_number = 1
        wat_number = 1
        resistance_number = 1
        tolerance_number = 1
        nlp = 1
        part_num = 1
        list_table_number = ["No", "REF NO"]

        pattern_kv = r"kv"
        pattern_v = r"(\d+(?:\.\d+)?)\s*(?:[vV])"
        patternv = r"(\d+(?:\.\d+)?)\s*(?:[kK]?[mM]?[vV])"

        for i in range(len(list_row)):
            data_item = data[list_row[i]]

            for s in range(len(data_item)):
                try:
                    something = re.search(patternv, data_item[s][0])
                    voltage_value = something.group(0)
                    for k in range(len(result_data)):
                        if result_data[k][0] == list_row[i]:
                            voltage_value = something.group(0)
                            matches = re.search(pattern_kv, voltage_value, re.IGNORECASE)
                            match = re.search(pattern_v, voltage_value, re.IGNORECASE)

                            if match:
                                voltage_number = 2
                                matches_data = re.findall(pattern_v, voltage_value)
                                result_data[k].append(float(matches_data[0]))
                            if matches:
                                voltage_number = 2
                                matches_num = re.findall(patternv, voltage_value, re.IGNORECASE)
                                result_data[k].append(float(matches_num[0]) * 1000)

                    break

                except:
                    pass

        if voltage_number == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")
            list_table_number.append("VOLTAGE")

        for i in range(len(list_row)):
            data_item = data[list_row[i]]

            for s in range(len(data_item)):
                try:
                    matches = re.findall(patternwat, data_item[s][0], re.IGNORECASE)
                    if matches:
                        wat_number = 2

                        combined_values = [f"{match[0]}{match[1].lower()}" for match in matches]
                        combined_result = " ".join(combined_values)
                        wat = combined_result
                        for k in range(len(result_data)):
                            patternkw = r"(\w+)\d*KW"
                            patternrw = r"(\w+)\d*W"

                            if result_data[k][0] == list_row[i]:
                                result_data[k].append(wat)
                        break
                    else:
                        wat = ""
                except:
                    wat = ""





        if wat_number == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")
            list_table_number.append("RATED_POWER[W]")





        for i in range(len(list_row)):
            data_item = data[list_row[i]]
            for s in range(len(data_item)):
                try:
                    tolerance_value = ""
                    match = re.search(pattern_tor, data_item[s][0])
                    if match:
                        tolerance_number = 2
                        tolerance_value = match.group(0)
                        # print(tolerance_value)
                        for k in range(len(result_data)):
                            if result_data[k][0] == list_row[i]:
                                result_data[k].append(tolerance_value)
                        break
                    else:
                        pattern_percent = r'(\d+)%'
                        match3 = re.search(pattern_percent, data_item[s][0])
                        if match3:
                            tolerance_number = 2
                            tolerance_value2 = match3.group(0)
                            for k in range(len(result_data)):
                                if result_data[k][0] == list_row[i]:
                                    result_data[k].append(tolerance_value2)
                            break

                except:
                    resistance_value = ""



        if tolerance_number == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")
            list_table_number.append("TOLERANCE")

        patternom = r"(?:,\s*)?(\d+(?:\.\d+)?)(?:\s*(?:㏀|Ω|k㏀|kΩ|mΩ|㏁|MΩ))\s*\*?\d?"
        k_pattern=r'(?<![A-Za-z\d])\d+K(?![A-Za-z\d])'
        print(result_data,"~~~!!!!!!!")

        for i in range(len(list_row)):
            data_item = data[list_row[i]]
            for s in range(len(data_item)):
                try:
                    database_list = hello2(db)

                    matchnorm = re.search(patternom, data_item[s][0])
                    match_k=re.findall(k_pattern,data_item[s][0])

                    print(data_item ,"@~")



                    if matchnorm and character=="R":
                        resistance_number = 2
                        resistance_value = matchnorm.group(0)
                        patternmega = r'㏁'
                        patternmili = r'mΩ'
                        # print("~", resistance_value)
                        for k in range(len(result_data)):
                            if result_data[k][0] == list_row[i]:
                                pattern = r"(?<!\w)(\d+)\s*Ω"
                                patterned = r'[KΩ|㏀]'
                                result_data[k].append(om(resistance_value))

                        break
                    if not matchnorm and match_k:
                        if character=="R":
                            resistance_number = 2
                            resistance_value = match_k[0]
                            for k in range(len(result_data)):
                                if result_data[k][0] == list_row[i]:
                                    result_data[k].append(extract_number(resistance_value))

                            break




                    else:
                        resistance_value = ""


                except:
                    resistance_value = ""

        if resistance_number == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")
            list_table_number.append("RESISTANCE")

        temp_number = 1



        for i in range(len(list_row)):
            data_item = data[list_row[i]]
            for s in range(len(data_item)):
                try:
                    tolerance_value = ""
                    match = re.search(patterntemp, data_item[s][0], re.IGNORECASE)
                    if match:
                        temp_number = 2
                        tmp_value = match.group(0)
                        for k in range(len(result_data)):
                            if result_data[k][0] == list_row[i]:
                                pattern = r'\d+'
                                number = re.findall(pattern, tmp_value)

                                result_data[k].append(float(number[0]))
                        break
                    else:
                        resistance_value = ""


                except:
                    resistance_value = ""

        if temp_number == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")

            list_table_number.append("TEMPERATURE")

        #
        nlp_number = 1
        for i in range(len(list_row)):
            data_item = data[list_row[i]]
            for s in range(len(data_item)):
                try:
                    tolerance_value = ""
                    match = re.search(patternnp, data_item[s][0], re.IGNORECASE)
                    if match:
                        nlp_number = 2
                        nlp_value = match.group(0)
                        # print(tolerance_value)
                        for k in range(len(result_data)):
                            if result_data[k][0] == list_row[i]:
                                result_data[k].append(nlp_value)
                        break
                    else:
                        resistance_value = ""


                except:
                    resistance_value = ""

        if nlp_number == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")

            list_table_number.append("CAPACITANCE")






        pattern_caps = r"(X7R|X5R|COG|NPO|X5S|X6S|C0G|X7T|X7S|Y5V|X6S|NPO)"

        Grade = 1
        for i in range(len(list_row)):
            data_item = data[list_row[i]]
            for s in range(len(data_item)):
                try:
                    match = re.search(pattern_caps, data_item[s][0], re.IGNORECASE)
                    if match:
                        Grade = 2
                        Grade_value = match.group(0)
                        for k in range(len(result_data)):
                            if result_data[k][0] == list_row[i]:
                                result_data[k].append(Grade_value)
                        break
                    else:
                        resistance_value = ""


                except:
                    resistance_value = ""

        if Grade == 2:
            max_columns = max(len(row) for row in result_data)
            for row in result_data:
                if len(row) < max_columns:
                    row.append("None")

            list_table_number.append("GRADE")


        package_number=1
        part_number=1

        if character=="R" or character=="C":
            package_number = 2
            for i in range(len(list_row)):
                data_item = data[list_row[i]]
                for k in range(len(result_data)):
                    if result_data[k][0] == list_row[i]:
                        strings = filterate(data_item)
                        # print(k, strings)
                        result_data[k].append(strings)





        if character=="L" or character=="D" or character=="Q":
            package_number=2
            part_number=2
            for i in range(len(list_row)):
                data_item=data[list_row[i]]
                for k in range(len(result_data)):
                    if result_data[k][0]==list_row[i]:
                        strings=get_Lpackage_filter_Lpack(data_item)
                        result_data[k].append(strings)

            for s in range(len(list_row)):
                data_item=data[list_row[s]]
                for k in range(len(result_data)):
                    if result_data[k][0]==list_row[s]:
                        strings2=get_Lpartnmuber_filter_Lpack(data_item)
                        result_data[k].append(strings2)
            if character == "L":
                for i in range(len(list_row)):
                    data_item = data[list_row[i]]
                    for s in range(len(data_item)):
                        try:
                            patternuH="(\d+(\.\d+)?)\s*uH"
                            match = re.search(patternuH, data_item[s][0], re.IGNORECASE)
                            if match:
                                uH_value = match.group(0)
                                for k in range(len(result_data)):
                                    if result_data[k][0] == list_row[i]:
                                        result_data[k].append(uH_value)
                                break
                            else:
                                uH_value = ""

                        except:
                            resistance_value = ""

        if character == "D":
            pattern = r'(?<![A-Za-z])\d+(\.\d+)?\s*(A|KA)(?![A-Za-z])'

            for i in range(len(list_row)):
                data_item = data[list_row[i]]
                for s in range(len(data_item)):
                    try:
                        match = re.search(pattern, data_item[s][0], re.IGNORECASE)
                        if match:
                            uH_value = match.group(0)
                            for k in range(len(result_data)):
                                if result_data[k][0] == list_row[i]:
                                    result_data[k].append(uH_value)
                            break
                        else:
                            uH_value = ""

                    except:
                        resistance_value = ""







        if package_number==2:
            if character=="C":
                list_table_number.append("PACKAGE")
            if character=="R":
                list_table_number.append("PACKAGE")




        if character=="L":
            list_table_number.append("PACKAGE")
            list_table_number.append("PARTNUMBER")
            list_table_number.append("uH_value")



        if character=="D":
            list_table_number.append("PACKAGE")
            list_table_number.append("PARTNUMBER")
            list_table_number.append("amphere")


        if character=="Q":
            list_table_number.append("PACKAGE")
            list_table_number.append("PARTNUMBER")


        print(result_data,"~~!!!!")
        print(list_table_number)


















        for row in result_data:
            if row[1].isdigit():
                row[1] = character + row[1]

        for row in result_data:
            try:
                num = len(character)
                row[0] = int(row[1][num:])

            except:
                row[0]=None





        result_data.insert(0, list_table_number)


        df = pd.DataFrame(result_data[1:], columns=result_data[0])


        print("@@",result_data)


        Q_table=['No','REF NO','PARTNUMBER','PACKAGE']
        A_table = ["No", "REF NO","PACKAGE" ,"RATED_POWER[W]", "TOLERANCE", "RESISTANCE"]
        B_table = ["No", "REF NO","PACKAGE","CAPACITANCE", "VOLTAGE", "GRADE", "TOLERANCE", "TEMPERATURE"]
        L_table=['No','REF NO','PARTNUMBER','PACKAGE','uH_value']
        D_table=['No','REF NO','PARTNUMBER','PACKAGE','VOLTAGE','amphere',"TEMPERATURE"]
        sorted_df = df.sort_values(by='No')

        # print("@@@", sorted_df)



        if character == "R":
            column_order = A_table
        if character == "C" or character=="EC":
            column_order = B_table


        if character=="L":
            column_order=L_table

        if character=="D":
            column_order=D_table


        if character=="Q":
            column_order=Q_table

        for column in column_order:
            if column not in sorted_df:
                sorted_df[column] = float("nan")  # 모든 값은 NaN으로 설정합니다.

        if character == "R":


            sorted_df.loc[(sorted_df['PACKAGE'] == '0402')&(sorted_df['RATED_POWER[W]']=="None"), 'RATED_POWER[W]'] = '0.031'
            sorted_df.loc[(sorted_df['PACKAGE'] == int('0402'))&(sorted_df['RATED_POWER[W]']=="None"), 'RATED_POWER[W]'] = '0.031'

            sorted_df.loc[(sorted_df['PACKAGE'] == '0603')&(sorted_df['RATED_POWER[W]']=="None"), 'RATED_POWER[W]'] = '0.05'
            sorted_df.loc[(sorted_df['PACKAGE'] == int('0603'))&(sorted_df['RATED_POWER[W]']=="None"), 'RATED_POWER[W]'] = '0.05'

            sorted_df.loc[(sorted_df['PACKAGE'] == '1005')&(sorted_df['RATED_POWER[W]']=="None"), 'RATED_POWER[W]'] = '0.0625'
            sorted_df.loc[(sorted_df['PACKAGE'] == int('1005'))&(sorted_df['RATED_POWER[W]']=="None"), 'RATED_POWER[W]'] = '0.0625'

            sorted_df.loc[(sorted_df['PACKAGE'] == '1608')&(sorted_df['RATED_POWER[W]']=="None"), 'RATED_POWER[W]'] = '0.1'
            sorted_df.loc[(sorted_df['PACKAGE'] == int('1608'))&(sorted_df['RATED_POWER[W]']=="None"), 'RATED_POWER[W]'] = '0.1'

            sorted_df.loc[(sorted_df['PACKAGE'] == '2012')&(sorted_df['RATED_POWER[W]']=="None"), 'RATED_POWER[W]'] = '0.125'
            sorted_df.loc[(sorted_df['PACKAGE'] == int('1608'))&(sorted_df['RATED_POWER[W]']=="None"), 'RATED_POWER[W]'] = '0.125'

            sorted_df.loc[(sorted_df['PACKAGE'] == '3216')&(sorted_df['RATED_POWER[W]']=="None"), 'RATED_POWER[W]'] = '0.25'
            sorted_df.loc[(sorted_df['PACKAGE'] == int('3216'))&(sorted_df['RATED_POWER[W]']=="None"), 'RATED_POWER[W]'] = '0.25'

            sorted_df.loc[(sorted_df['PACKAGE'] == '3225')&(sorted_df['RATED_POWER[W]']=="None"), 'RATED_POWER[W]'] = '0.33'
            sorted_df.loc[(sorted_df['PACKAGE'] == int('3225'))&(sorted_df['RATED_POWER[W]']=="None"), 'RATED_POWER[W]'] = '0.33'

            sorted_df.loc[(sorted_df['PACKAGE'] == '5025')&(sorted_df['RATED_POWER[W]']=="None"), 'RATED_POWER[W]'] = '0.66'
            sorted_df.loc[(sorted_df['PACKAGE'] == int('5025'))&(sorted_df['RATED_POWER[W]']=="None"), 'RATED_POWER[W]'] = '0.66'

            sorted_df.loc[(sorted_df['PACKAGE'] == '6432')&(sorted_df['RATED_POWER[W]']=="None"), 'RATED_POWER[W]'] = '1'
            sorted_df.loc[(sorted_df['PACKAGE'] == int('6432'))&(sorted_df['RATED_POWER[W]']=="None"), 'RATED_POWER[W]'] = '1'

            sorted_df.loc[sorted_df['TOLERANCE'] == 'F', 'TOLERANCE'] = "1%"
            sorted_df.loc[sorted_df['TOLERANCE'] == 'M', 'TOLERANCE'] = "20%"
            sorted_df.loc[sorted_df['TOLERANCE'] == 'J', 'TOLERANCE'] = "5%"
            sorted_df.loc[sorted_df['TOLERANCE'] == 'K', 'TOLERANCE'] = "5%"

        if character == "C" or character=="EC":
            sorted_df.loc[(sorted_df['GRADE'] == 'Y5V') & (sorted_df['TOLERANCE'] == 'Z'), "TOLERANCE"] = "+80,-20%"
            sorted_df.loc[(sorted_df['GRADE'] == 'X7R') & (sorted_df['TOLERANCE'] == 'K'), "TOLERANCE"] = "+10,-10%"
            sorted_df.loc[(sorted_df['GRADE'] == 'X5R') & (sorted_df['TOLERANCE'] == 'K'), "TOLERANCE"] = "+10,-10%"
            sorted_df.loc[(sorted_df['GRADE'] == 'NPO') & (sorted_df['TOLERANCE'] == 'F'), "TOLERANCE"] = "1%,-1%"
            sorted_df.loc[(sorted_df['GRADE'] == 'NPO') & (sorted_df['TOLERANCE'] == 'G'), "TOLERANCE"] = "2%,-2%"
            sorted_df.loc[(sorted_df['GRADE'] == 'NPO') & (sorted_df['TOLERANCE'] == 'J'), "TOLERANCE"] = "5%,-5%"

        else:
            pass

        if character=="C":
            grade_temperature_mapping = {
                'X7R': '+125°C',
                'X5R': '+85°C',
                'X7S': '+125°C',
                'X7T': '+125°C',
                'C0G': '+125°C',
                'Y5V': '+125°C',
                'X6S': '+105°C'
            }

            # 'GRADE' 열 값을 기반으로 'TEMPERATURE' 열 업데이트
            sorted_df['TEMPERATURE'] = sorted_df['GRADE'].map(grade_temperature_mapping)


            # print(sorted_df)
        sorted_df_by_column_order = sorted_df[column_order]


        charact_sheet = work[character]
        for row in dataframe_to_rows(sorted_df_by_column_order, index=False, header=True):
            charact_sheet.append(row)

        last_row = charact_sheet.max_row

        for i in range(last_row):
            charact_sheet.cell(row=i+2, column=1, value=i+1)



        if character=="Q":
            openai.api_key = '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'
            question = "is it FET? or BJT? just say FET or BJT."

            FET = []
            BJT = []
            ELSEQ = []
            for i in range(len(result_data)):
                string = result_data[i]
                len_last=len(string)-1
                part_number=string[len_last]
                response = openai.ChatCompletion.create(
                    model="gpt-4",  # 'text-davinci-004' 대신 최신 모델 사용
                    messages=[
                        {"role": "system", "content": "You are a helpful assistant."},
                        {"role": "user", "content": "partnumber is" + str(part_number) + question}
                    ]
                )

                if response['choices'][0]['message']['content'] == "FET":
                    FET.append(string)

                if response['choices'][0]['message']['content'] == "BJT":
                    BJT.append(string)

                if response['choices'][0]['message']['content'] != "BJT" and response['choices'][0]['message'][
                    'content'] != "FET":
                    ELSEQ.append(string)

            head_list=['No', 'REF NO','VOLTAGE','PACKAGE', 'PARTNUMBER']

            FET.insert(0,head_list)
            BJT.insert(0,head_list)

            print("안녕")
            print("@@",FET)

            print(BJT)
            work.create_sheet(title="FET")
            work.create_sheet(title="BJT")
            work.create_sheet(title="해당안됨")
            charact_sheet1 = work["FET"]
            charact_sheet2 = work["BJT"]
            charact_sheet3 = work["해당안됨"]

            FET_df = pd.DataFrame(FET)
            BJT_df=pd.DataFrame(BJT)
            ELSEQ_df=pd.DataFrame(ELSEQ)


            print(FET_df)
            for row in dataframe_to_rows(FET_df, index=False, header=False):
                charact_sheet1.append(row)

            last_row = charact_sheet1.max_row
            for i in range(last_row):
                charact_sheet1.cell(row=i + 2, column=1, value=i + 1)


            for row in dataframe_to_rows(BJT_df, index=False, header=False):
                charact_sheet2.append(row)

            last_row = charact_sheet2.max_row
            for i in range(last_row):
                charact_sheet1.cell(row=i + 2, column=1, value=i + 1)


            for row in dataframe_to_rows(ELSEQ_df, index=False, header=False):
                charact_sheet3.append(row)

            last_row = charact_sheet3.max_row
            for i in range(last_row):
                charact_sheet3.cell(row=i + 2, column=1, value=i + 1)





    work.save(output_excel)

    # BytesIO의 파일 포인터를 처음으로 이동시킴
    output_excel.seek(0)

    return StreamingResponse(output_excel,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=result.xlsx"})


@appends.post("/upload/")
async def upload_excel_file(request: Request, file: UploadFile = File(...)):
    file_data = []

    # 임시 파일로 저장
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        temp_file.write(file.file.read())
        temp_file_path = temp_file.name

    # 임시 파일에서 엑셀 파일 읽기
    workbook = openpyxl.load_workbook(temp_file_path, data_only=True)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        file_data.append(row)

    # datetime 객체를 문자열로 변환하는 함수 정의
    def default_converter(o):
        if isinstance(o, datetime):
            return o.__str__()

    # JSON으로 변환
    encoded_file_data = json.dumps(file_data, default=default_converter)

    # 임시 파일 삭제
    os.remove(temp_file_path)

    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "file_data": file_data,
            "encoded_data": encoded_file_data,
            "file_path": file.filename
        }
    )
